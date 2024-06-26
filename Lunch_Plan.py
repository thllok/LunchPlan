from turtle import *
import turtle
from random import randint
import time
import PySimpleGUI as psg
import random
import sys
import os
import base64
import PySimpleGUI as sg
import pandas as pd
from datetime import datetime,timedelta
from openpyxl import load_workbook
restaurant=load_workbook("restaurant_list.xlsx",data_only=True)
layout = [
    [sg.Column([
        [sg.Listbox(restaurant.sheetnames, size=(20, 6), key='-LISTBOX1-')]
    ]), sg.Column([
        [sg.Listbox(["正常/糧尾","公司請"], size=(20, 6), key='-LISTBOX2-')]
    ])],
    [sg.Button('Have In'), sg.Button('Take Away')]
]
window = sg.Window('Two ListBoxes Example', layout).read(close=True)
restaurant=pd.read_excel("restaurant_list.xlsx",sheet_name=window[1]["-LISTBOX1-"][0],header=0)
restaurant=pd.concat([restaurant[(restaurant["Budget"]==window[1]["-LISTBOX2-"][0]) & (restaurant["Have In/Take Away"]==window[0])],restaurant[(restaurant["Budget"]==window[1]["-LISTBOX2-"][0]) & (restaurant["Have In/Take Away"]=="Have In/Take Away")]])
monday= (datetime.today() - timedelta(days=datetime.today().weekday())).replace(hour=0,minute=0,second=0,microsecond=0)
restaurant_tried = restaurant[restaurant["上次食係幾時"].notna()]
restaurant = pd.concat([restaurant_tried[restaurant_tried["上次食係幾時"]<monday],restaurant[restaurant["上次食係幾時"].notna()==0]])
restaurant = list(restaurant["餐廳"])
img=base64.b64decode(b'R0lGODlhMgAyAPcAAAYDAAsFAA8IAQwMDBMKAR0PAR4QABMTExsbGyAPACMTASkWAS0ZATMcATkfAS4gCj0hASMjIywsLDExMTw8PEMlAUgnAUwpAVAlAVMtAVkrAVcwAVwyAWMuAWM2AWc4AWw7AXM+AXZAAXtDAUNDQ0tLS1RUVFtbW2RkZGtra3R0dHx8fISEhIyMjJSUlJubm6Ojo6ysrLOzs7y8vMPDw8zMzNPT09vb2+Pj4+zs7PT09P///wsFARIJARISEhwcHCMTAC4ZATEXATs7O0EeAUknAUQpC1MnAVosAVIwCXI+AUxMTFNTU1xcXHJycpycnKurq8TExMvLy+vr6wUDABEJARQUFCkWAC8dCDIXATwlCSQkJEkmAUcoBEwqAVImAVwrAVsyAVg1C0RERFVVVWxsbIKCgpOTk7S0tLu7u9TU1Nzc3AsLCyEPASsrK0IfAUYoBVpaWnNzc4uLiyoWAS8gDUMfAUQlAVorAWA5DIODg6SkpDQcATkkClssAVExDGY4AWNjY3t7ewsGABAHABcRChsSCh8YECQZDS4fDyYcEi8iFDgpGVQtAUQfAUgqCG1tbfPz8xgQCCEWCiIYDTYgBlI2FEM2Jx4WD00vDlElAXZBAb29vSoWACUWCikaCDUiC0MkAVQuAeTk5BsUDCAWCicbDSobCjciCTMmFjIyMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACH/C05FVFNDQVBFMi4wAwEAAAAh+QQMCgD/ACH+KEdJRiBlZGl0ZWQgd2l0aCBodHRwczovL2V6Z2lmLmNvbS9yZXNpemUALAAAAAAyADIAAAj/AHcIHEiwoMGDCBMqXMiwocOHECNKnEixosWLGDM6lKKx4ho0L6B0jDjlhholHADtGekwTZMCBRpgCAFoTg2WCNGQIQCzwAUlIUaM2PTBBU6DOmi4KcAHhNCnI/SIkXFUh4w9K7YUMHBhk1ARUPM0ShEJZw5OS2C28QDV6VChXjdxEqhjxw2Me7QayOJWKFAlI8A+BeEAxY0CKlqYyYFwyijGCHXsmQCTSwjBI0B4DaE5cGBRBqoMMkC6gJsXNUbVzQFjE9gQKtDgMJhG0BADHKCOUOJVqI1Ra6LM4RCEgIBBW7eWhpmlt+5NsW1M2aFj1J48gKGKyD7CKMFInKoI//ih3EDPAgu464a6KUTQ9fAhF7wx2zxMSQWayHAOv79/oVQphMIcNEhRw3Rq8KdbX/891VtdEb3w33YNsjcCGhMJ0iCDDYIlkkQtbFjhYCOsJJENIo642wgxUPSEf+6puGIUFYW4XgjqrZejZvJJpAMMuuGo4ILagZBCFGuURZEZg4VAwyho9MfhBx88pQQISugx3UNTDOZVGnRx8haJQynhASCb9Lada2A+dMZufbU50BoynCEHUb0pAQUMc6ywggto2AAJCBw5ZEMeHkCC3ZcLySDYhwW5wIecDMnQQBAKZOqAKDQwBANYIsxW0BM9BOjQKFJA0QQBrKJA40IugMnFgkEyVAEDRYEUQEAPBJClkA4sgDUXQWj04J1EMrRRQA8CCGDGQjlU+SxBMQgwR0U2kNBsszcNpEOSAuWwB5oYEjTFECpYlAMKAlhRSAsC1cBEHCXsscYZ6EFSaLhPLCGAiRVFAskWgQQyUBpDEFBFFUPAsIZBU6hSSBWjYKSDE4WosuUOOcTQQhoQQozCBHpopIMKBIwhakM6TODGyhbboIoqDzvESc0dTfHDFiQoeVRDL6AwRLc/N4SDGaYWrfTSTDft9NMTBQQAIfkEDAoA/wAsAAAAADIAMgAACP8AdwgcSLCgwYMIEypcyLChw4cQI0qcSLGixYsYM2rcqDCSGhg6OD5cY2PKlEB5IE0RqVBHGhkyTEgqQKfSBxBQbLAs6FJKFBUFCoQCMaJoURBodg60AUNR0AUejBrdlAQEDqU7bqSYaUGEVKMhxBTVKXLUkzgFFEQ1KkLJ16JKYmzUQeMMhQIQ3IINMYLoVw90zMhAI2UlRR1qZLjY4FeqXxFElYBQIsJLgSpBgzohO/BGDDMqVjyREslgDiiW+L5tvEnNlJA3XjggIMCAgcwFUMSQombF2xGQQQifrPr3iE1GXxi8EaF2Adu2M38wTr260dIGbdzAHTQQiyZvKVv/r76J80FJS+KcURNyh3Elesd/DZFm4Zod2AXe8PpbvHzHhkE0A3+/NfYfCJtAIREnBn7V4HgghCCIRDhMZ9yD1glnxkRmYNhXhg6CsMdENsDXH4hSfaAEDaOoIUUanDw0h4kOoljUJh5sotcmGCwghUNTiADZV0+oAUUKyK3GVo5SAXJJDzE6NMdxBt4w0Cg0uDACfG7JUVxfIKSQUyQ5uKHCQzEcpVcKCbW3gw6cSLVJfQO9IECAC8UJV1F0NuSSVCrct4MaVfS5UBRFhcBffg7RVYYXobggkA6qQOIQFEcd9wEUjDpEAwFV/CiQIAJ0itCNMqjAQSUFDJGGmwzp0rBEFZIOtIYAgipkwyYi0CDQGjKUIYAATIjK0AsEQOKmDiz0oAZDaWRwx1UEjfICKVWUkQNCITFbRRzb2joEAdQqNIUqBIxo0CglCKCsQToIwgIMPUAS7kA6BKIKrAnZIMC+B0VyRhVPFDRKCib00MMe/Ar0gnIO7VFFlAaZ0UMB5e4QxRlOQIGnwffGOgQZCO0RxxjqYrVDDALUgNAUnCSl8g45CNBEwzMb1GzKOSN0AxlVvDBKzwlFgsAQWzxL9EGqkFFCE0sjFIlJUVdtNUIBAQAh+QQMCgD/ACwAAAAAMgAyAAAI/wB3CBxIsKDBgwgTKlzIsKHDhxAjSpxIsaLFixgzOtQxJZJGiTnS5FCzSYXHjw/XQNkSKBQSQVN0oDyIA8eOKGeGFDBQgM8GCCJCKGkxZSbBGoLGMJFUoICDDyOiRhURSMnJmTpuNCEgyQCdPFLDhijjwYXRSDRIMDWQQUTYtyCixviIA80TM0wtKIEb9+0IQFAyrpmRpkWDMHvhjtjUV2qIOwQiwdhzwyIOF2D9jggaFcQKJUoEbcqjoEpTAgVYFB2og0aLFCBAKJHDAkYMTjRooDmjOWxfJ6MK2igjYKcBngX0oJFio0ni3tChrzZohufx6zwVPI/OPWwNhU2bSv9qskdKCg/d00s1q1DSmBJpVqtprL57lIU3IskcGIO+VBCb1CfVJqFJBIN//wko1QcgTOfQgdAh2F1cakQUA2eaSfiWW2Ht9R1ENoygoYZhibBdXJxIRCCHbxE4oW8jPLHGVQ2pICKLYaHHXWOb6LjZCCzssZ9CIUSlRJFv1TBKDTKcUYYSHLpo44CxxbZJEQXssVANYvlng0E5TCGmTJHwhqSRHFgyAwVuOFiQDmWE5UIOgoT15UZrSNHkCQVQsIMNVQiCEA5mDCiDQDmEVeFEowgQiEB7EICGQS8EuBgUORwlVWUTAcqCQDqQ0QOnAyWGBo0DpRBVcBNFUcVcAknd0UMcBYXywYcHwbAqRWhUQeqfW6xBUCAL3HmQDLtKpAMkJRR0QxXG7sCCALgaFAUIHKQo0Rw9TEqQCkvMQNALVVRrEAzuDdlQJDa4EAGqN0Ci5UAwCBDtQS4I8AREUQRSBQwHqTGvQDJQy9AKAvyK3wlLZMpQGgLQwJAOTcShrkI6xCFsQ5wIkEZDo0TQglE70CAAwA1BUQULqGI0ChP6sqpQDj8IQAINF1+kA8yHaDuzDQJUocq9FzkRSByqtHzQHCSkQMLGGU1xgxo2NbRGGpxwAjXJXHft9dcDBQQAIfkEDAoA/wAsAAAAADIAMgAACP8AdwgcSLCgwYMIEypcyLChw4cQI0qcSLGixYsYM0LUcSOHxoo6xAjS8fFhjjIo5kBp8gFGSYVTIt1Io6iAgQKhOHwYoeKMlJcGR7UgYDNIGBEjkir1wIIk0ClO5hDNkkep1aQpNkT5qCPKExYsSBRo4OGqWQ+b5mgcxYkGjTkKuBy1ugnEJiVmR3y8IUfShU1mQeAdAcKskh051ESyGEnNisKBkSpFsSmE4BEbXsSA4QTFHhwDdaRxIYiFmRZPoHCqYaPGDM4h8iYVMTipC5IecUARe7NAzQJmatxIoQSy7ONWa49Ac1CHgefQoReopBy5daUyEkYvwD1QqCKSr4v/H2FGIfcCqpZEaVx3vHgRINSYl7RHsUAoxt1f35pQzZSC+Mm2CWD6KdVCRDPkd5WC7q0QUQ2EycaghFepBZEaEeY1YV5KEJhUDBHZMAJtHLqX3wwRzZGUXYGtUB2FVkESAw01yKfQHlYxKMgOOqiBhgsqCBaCEkSq4EIUMcimhAc1IKTDC/s1F4lToa1hwx4gZAmCC0y04ZJBUwhiFg44KsUCRSImxckoApxZECdm7eGRFFZ5RNEUKoyghBo/oFDQKBbg5YQMNxBEw2yFWqSDCx6UMMYPBcmQABk2HJTnCGloFAMBAghA5Q4mEACaQTMk9UJJUrghwE8C1VDFgQe5jgCIB6N+JIgAP1S6QwoCJGpQFAsogGJJNQhQhRvyCRCIQoFU0URJkcwhgAo9RCBFFZwodEOnvmJUQyAUzHBGDwRQ8OlBLwhwRkktDJHpFGtkt5AOSwhgZ0Y6wLDGQPcqZIMAUAAlUSBudCswQ2o4UYUZ+x7sEAyYDEGDww1JQUIKQ/RLMUJTbuzxxyBTHBAAIfkEDAoA/wAsAAAAADIAMgAACP8AdwgcSLCgwYMIEypcyLChw4cQI0qcSLGixYsYM2rcyLEjQh03PC6UgiMSpxcuoDyBNEeHSIM6bECJMAFSoAILHDgIA6nGDZcvR6VxUaBogiIfRij1oHREmY44RulYE2VOUSIcRDTdqjSPxzQEmIyBwCEEV6VmlYJ41HFUHDob0p4FoQTEVhEOgS7UoeLsVhB2RwRuKjDKmoNqWuTxAELFGRgx0nCiUcOGDSky+vodUffvX4E2nKh6EVIgjM6bU/vd1HSTjEg6onCqgYaFKgMGihYo8ULQBiWqg6c+bBC38eMJBgtfPmIT8eLHi5rhwJT5chAi0CTEXVQSi1E1BFv/Z+5Br8ECBFi4UCPQBWe/wMdztZFwikGzypvmly/j4Sil8XG1n2pyKfXEQ1CoBd94ILDW1IENceLggAOqppx2e8Hg2VkeKOHgda3JgENCOkjhhIBnQSLQFDag4Z5Wm23yQYAAzpGGfQPRcAkgXG2y34gGRWKDDC2oAMkKT3BymAyAAQbjCB4IEcNAFLRhxBMxzNFCGqO4txWQEp3RVBo30ABDIARsAZQkQ5Q2UCQfpjBKRTqcOMIKBKFBCA0CDfHcQC80dUYkF/2n1J9MpCDQnzvo4EJ8fGYkxSZKQCjQDASASRAnDlggx5wbvRAKIzkMpIMbMByEBgGSEMqRDkwUmNDfQCqoYp5AUggggEg29DDGpgREYVAOur40Rw/CDsSECgc1IYCrHU3xg6IC4XAIAQftQQCOHkFRBX07jOKEigbVQACoHulAARmu6vDCQVMQwt5LnBCAArcHETvvSzKoQoGbBq0hwAqlvrTDHCSogq5BKFShihQG15BCCnMkVAMJKEgA7ks3wLAwQSZxgsa+Bpds8skoJxQQACH5BAwKAP8ALAAAAAAyADIAAAj/AHcIHEiwoMGDCBMqXMiwocOHECNKnEixosWLGDNq3Mixo8eMNzjtyJGGho6PCXWskTIDEgkzqTKkwIGSoA4pNFAskVTAAIQNHxplUKIkBUodU2rQKEPAgAVAm0ZIHRF1xImPUghoLbBgQ4ipYKcy8SglzYsClTJUDSsCxFQRHKekaXEpj4iwU5W4dTtVxw0bakYZ1EHjyRMXZwxD4WQDR6SBkabYYIEXLIiqfKWiOQOJiSRFKvZIkRwoc2URRJV8rWwZbAwakESE6Gmgdu2ePU2z3o1XSdUzOQbmiGSbtoECJyxs4M189yYaB20nKEBhiY0oIe423y71rsiDBcIT/2Ah5bEZ3dx5b0KRcEsLGt933BiBPr3zxwdPFqRcf8Rq+2BJ4VAUI/jGWn/cwdBQDVL1hyB3KygUiRSUNViZEtrZZ5ogCNnQRB5hoYcagFKF8N8cCEnQRihKTPVBZVKoIaMaUsiwQoa8vSgVJAhV0UQUNKSRRgxrSQUcQjnUEANnGEq1iW8g6LVWfARBot8Oo4QliBoQ5TDFKFPskMZUKlS1CZUC4ScfWCrUcNEUVUUxRQtTydDhZSG04GZGgkgl4A40TAXFQXI4oEJwG8kglQ0DjaJCCBnYWZAbBAjGkRqL2nRGAVWsQdAUVXDY0RQcfMBlQTEIEMeVaQgAnUcEtIhxakGcCLCCfpBUgWhHcQhA00FSkMIClgWw95EgAlh6UA2F7CADAYN+5IIAYSY0aBNVzNoRFFVUixAOUwgQwZUdodHtQpFgMixKnFTxa0I3CKBCTTf0sIKaB52higR7YrXEEGiQS1AOKaDQQk07qPFDCZDsWtAUT7yA8CiCwPCCtghnrPHGCwUEACH5BAwKAP8ALAAAAAAyADIAAAj/AHcIHEiwoMGDCBMqXMiwocOHECNKnEixosWLGDNq3Mixo8ePOD42tEFSjQtBgkaJPKiDRhQaZ1QVaAMhA4gYKwmmeYGCSYEgG0CMGCF0xEodUs6kWFKgQCUPQ6NGXelyBooCDgBJ3Wp0h5o0aGZIsYFDB8FINWTAkHFj4Y05KUq8sclViRIRUU803VvAQNMlc2TMwSsV0gyVBte0aKCV61AQmx4PbWrAb2W/lIs6HqFkhCAzoM0IiryZqNRAQy9fbkpmBQ2opWOX1syi7Y41lpuyaMJJxo45pGULl0p6T8GmQ6Dk4BRpyo4bmodLHwrFYAwbZgu2IDx9eIgzD3U0/+4+/INzh2qiEye/VYlCHTGkqPlqmqt69iNCIpSyNQR3qfcJ998INiTUwgikbRJggMKpV+BBkXSmhiBKBCiIDDHs8QKF0okQglSIGcTJBnNEwUJwUYmwR3YD6XADFMNp5h5CkNABG3EpyHBeQjjQ8AKCjmkGAkJrCFCAUHpkuAcNIT6UgxpmgCDlE2tMAclQSrBIEBNVqJLBExo9MdQmiNlAmhoHVcEEASVoadGPI6xAUAxDgWnQDywQwMlGcwwlRUF9irDjQGkswQRHTgzlZiQjhPCCQTVUsadGOYTxARoHqREGH2sURAYJblpEQxup5IDQHgYIUlAPmG7kQhWPJmYESRU1ELREqBYFUoV+CE0hARMsmrqRDgKowFAUAhgHUhVRNNSCALZ1xEkVwi4UiSpLVLvRCgI89IQAS9TKUQQCzPBQUlV0pFixDzI0BQUfkSHBEHYylMZHMKAQSCBo5uTvvwBXFBAAIfkEDAoA/wAsAAAAADIAMgAACP8AdwgcSLCgwYMIEypcyLChw4cQI0qcSLGixYsYM2rcyLGjx41TBqphIeWjQh17WKhycsYJBxAsTB5U88IMkwIF7njI0GiETII65hAoYKANBiUjkoLwKVNHDpSBJCnYICKp1aQDc6hxccbFHk44CK6J4QJGlFENR6kIpMoAhqpXN4lAuoLMUJw4DeBcEuPGmU1x5dDQkRCNigIJiiC9OkLJ0qVEDeiVTFkynaWMr+6xEVKgjjRzLmEABDgzCLggQlSmjNeFmsywR2xSokSOijIhYidVAvcMjkircQqCEmPHGt3Ik49YuinKQKIFSpChsUIHGoFRlGtnjFkNwRQxpqz/2RGps8A029Mn3XQD4h716TkxJDywBWzM8K0qoZ9whhIWMtRgQ2lXiZBbfkq5wBB39yFoFQjFKYTDAigoRaB+cKWX4Qgi1CBhHI2g4AR+cdEQxV/pkbiJeQfJcJpuTrTnmRrvKUfifgqhoMAYMMwxIm2buGBDQmtAgRyJcyh0QxUCJAnUQza0sKFSV6WhUAuECHCdRWt8AMKXIEAiCIEyHpRDlgJ4Z9ETVs1BnxNJiRBJQnv0AIkAc1akRmlm8MeJVYVhYkIcTVikAwpICcIfeYDFhNAPEpDxwxMWyVDJHSjkSdBfTiREiBQtLDEkRZFIAMQlaBmUxgUZFLaDCmdYf5QGAYqoaZAaRC20x6ISNUFASQhN0YMEP+EgAKUJ6SBACj+9wASvBlXh5EdPEDKqQmMo+BEnhKzQkAAvePTCEyoIENZCZ06rkQ5MEEBAhAstGUGqGuWQQhXhOkRDFSR4qNEUElgJEQw/LMECtBPpcO1DUlCgAgXw/iTxxBQrFBAAIfkEDAoA/wAsAAAAADIAMgAACP8AdwgcSLCgwYMIEypcyLChw4cQI0qcSLGixYsYM2rcyLEjQh04BKpJE8mjQxguOEECAcIkwxp6ChRw4EGES4M4akBhUaIAHw4hRoywafAGjikcdcSRVCCIB6FQRwicUkOGCxWYZBZwwknHQB2jpnh1eMMMCTcFLoiICtXGEqYyDRQwIFduiSg25igRqsRMDYZqBKWQpOApWxBrtdKtK1drASRsoaagMZbglDQsnqAoAGFv1BAghIJYzFgmCjI3dpyJzHbTmSghBa6RMefNh6BsRXje1GIu3blOCkhR83fKJtbIlQhy8eSFE+QjQo9wkjqFnC1umtjIcUZgySjQw0f/hy79zNhIe6DsKbmDvcA04uOz3juHoI1R7RHOYB3iuHzo7jUkBWvS/ceaDQ7pEIUOahBoYG4jpNDQGjpwIhQKDj4IVWhQMBSDKEqkwBJyBWoYWhQMmbHAB+FBwsJaGo4Xm0JkEADDJv6xpUIOO+Qgg3yeiSYVQwIIEAkanSkRQghKzLEgQZH8GF6Q0aHBUA4CULADDASo0R6PCR2J3CZBgmBGZQnhUMUKO8xRhZcJSmFGjuNB1QKaCdlQRYcpVJFaRKOgIQhLLArFpkNnEPAXClXMGBEMQgr1J0NqVsFjDT3gJ1EOx4kAiVAqNERDGk9UYcYOOuxBAFISxSAUGmgItMUJQzqYEEEVKHw3RBWsQjSKUC3sAF8IASJUAyQCnOmdE5iA+ZANY4AASUkveHAqQ08skUZBZqAAUSQu9KBAHpqmoEAMDd2A5w41/KXQFCXlEAMFkrCgrncCrLDuRjiggAInaFFXEA0CwGlSDifUSEAg7hYkSCA37aCGCy5KgdAUBMwaMUN7uLHvxgTpoIoLIC+kJwwlKxQJGT3o4WzKBalBQRyBFAtze5xwgsakN/fsM8gBAQAh+QQMCgD/ACwAAAAAMgAyAAAI/wB3CBxIsKDBgwgTKlzIsKHDhxAjSpxIsaLFixgzatwoUUcOThwhRuIEhRMIJYJCLtQh0AYaEwX4gBihUqGOMygELSlQwIsSmjUR3nChimeomSOA1rCh5saNMy9kSMGxkZOZQwWyAEo6IsQIHWpa8BzL00CBMVByCJwSJcYeGpEaplGDAw0FA424jtg0s2gBswYABw5cQNETM5v0jkjD0mYMRQWCIE0aQkmIEIIH/zVbAJILUYoVy4hrMAcNM5IKWBChV8TPEZDGDjZrBoqiKDuihA69ac+axjqk7GnCB8OH0EjNREo9lkyauEsFrthNXckmFZDyeJmsODnLUVNo1P8YRZC8QO7Ud7NOj1Stw8S74aef/5MGxDPU0c/fvSKilPz7zbfJGhFFot8IBwaY1BkFvnBgggrW4NAegjzxxEm7QRigDQ3pQIABRXi1Wxm67bceV0ip4ZAqBCyRIFU5vCCfYpXpZd0IIrjHkBwETDHKCl5wcJIenDQm0BrTIdddUnM89EIVasnQgww6HiRFCnqhpwRrm1DlEBRVCFRDFWYwpAMNKiTGHQhebSIFRDD0sBYBEzw0ShqTgcDaHlNE1AITA6nSg5cN4YBiCCgY+VAOqkAykBlV2OdQGkkhhZtEowjAoJhLvPCQIF29NtEUMAiQBkFpkOEQDkdtNQJIEkmcUIUA5u2AwxYCkLbQCwQwAuomtUKEwhBOEJRDCqrcgBAOjY3ygwACBPFBShPNoQqBBMFwbWkwMDHHDS0MIoC4BJyha0QveFrQGiUYpMMSS5BBwKwCpKCGohNVSVCfB+VQAxppwKtvUAjlAIkbhBKckA4mCCChwgtFQgGgEC+Uw6znVmzQKFscgobGCkGRwh4gK0SDsiWnrPLKBwUEACH5BAwKAP8ALAAAAAAyADIAAAj/AHcIHEiwoMGDCBMqXMiwocOHECNKnEixosWLGDNWzKFx4o01O3DoeaGjY8Q5LnCcUZLihsmGo2AEKlDAAYgRIl4q1AEFkqQCF26OGKoTIZQnT1QxEDqUqMAcN2zYwCFDSqSOOlIUgBCiKdM1T0rQNFCAbAFFL0YV1HG1YaQoafY0KVBkk9cPTcca2FuWL9kWOHZIERQixCYXNhZGwrFHVYENTUeAsIt3xN7LmPuadRK5aYoZHA9KGVMgSOWhSkR4PkOWb9kCQ5iUoGFDhofOnUWoiFFDrUAakiAAws10xVUULOZIahIjRos1gto+wU29qRIVLlAMJ95UTtuQewLv/6CxI0fJ8qqrqx+hpLrQTb4b3nC/Xv2m9iPQQIyUPjf++p0JBcl5D81BH4DVRSGRDdX9h2BkIXwHESTUidDVg5FBQpEU1ClhF4bsjfACRKNIgcZ01DGlnooWjgAFRB+qp8YLMeLWoldD1QDRChzk4UR/Q+0R2hosHIjjCAQ2pAZNOUyhRB55xEeQGoJ0llpkk40ghUQrVEGeFJUs4BJCarjwoYohtLfHRDgIsIJALFQxx0I6rMEJhah5AAIIiU30hABX5aBKFSAxBMNQKjzxIXkT6TAGozT0EEiSCO0xFA06NJVGRS+gMFCXMDBk6SZT7MCgixWVUYWCO0wxhAB9JqxkqYYCGfhERU8QwOgOoyxBBqUFnTFCDAJNUQYIt040yiEFjCnQElVwktANeYSQWBqXKICCeAuqYEhoO+igSiCCIMTJFgrIgIMTBPwQA7APTcFCuVOaoQqlOcRZCBRrKEKAGaVeNIcMBukgbatrTKGHKj0wEUgVApChY0Y5SGjQGZjQNEYaMQjSAqtFGTRKCwQU4ILFIR8EwxAqOJuyQua9LPPMNNds8804YxQQACH5BAwKAP8ALAAAAAAyADIAAAj/AHcIHEiwoMGDCBMqXMiwocOHECNKnEixosWLGDNW1DFF48YnJ6R4hCiFRRkZNUZ8gDLSYQ4pLQo4CKNEScuFN/akKEDnw4ifI24mzAHDjCI+Sn6G+Cn0oI4XYwpAEPETBNAdkUZJsYFDRw01OUaucWKgCFCVVEcEKsC27RI3BEjMQYODoA4paGbc0KFwig0ZKgoYQHLWqtURBhIbEKx4ceICTmrsoHF2xKYWUvgWVCOlhqACDHwW/qkEBGPGigUnSJ2ncmVOmgXmQLNEQYa0QJVssnxj8pwhbRezdVNgjIsnrV27fjGDBg0XGUS7PgyD4JQnKaIWMCNFBQ2Ba3Yr/x9PHujhSAZz2LARtkZsKMo3LS1PPmmaiE/GgxBPf3xsh/Dp199430WEA3+VHTZgYehJtAd5Cg6YFAj/PaQDJAIuOIJVITQoUUqu6VYeghsyRZEMygVCX4QinjHRU5WVpsYON7Aw3iZJmfdTFBLlIAhpLGyynxwE2fBjgmeJCMIco0RUQ2kooOfECCiEVZANKwCF42ggULUJGhGlsUAQTujw4Af3HYQDDXu0IJ4SgcQwR2UreNiQGhQQkEIOMjhQQHUKxVDVQDacMUIIIIDgIkRTQNJDIDmoMUYPKKyREA6kGRSJEwu4IJEOZ1RhwhSRrFBFFU9UOFCmBY1CwBZ2QqdkBgEoCCQDAY9aWdALI9hUEAxVADpRJIJUYSmNnFSxREcFpYFjQTq4QYquE+1RhQwDRVIIIYEY9ER0BUVRxQsW6TDGD+QO1AIkbhy7QxQoEFIAmATNgAmzFbEgiBkErQDJCZ5KwcSre+iqgyBOkLFHRlA4YV0Mea5AgBtQULuDGkvQqqpHo5wRxxOxCsQJGU1Y3FRCUaSx8ckst+zyyzDHLPPMNGMUEAA7')
file = open('horse.gif',"wb")
file.write(img)
file.close()
number_restaurant = len(restaurant)

if number_restaurant<4:
    layout = [
        [sg.Text('得返'+str(number_restaurant)+'間餐廳今個星期未食 跑住'+str(number_restaurant)+'間先好無')],
        [sg.Button('好過無啦'), sg.Button('唔制')]
    ]
    window2 = sg.Window('', layout).read(close=True)
    if window2[0] == "唔制":
        sys.exit()
else:
    restaurant=random.sample(restaurant,4)
    number_restaurant=4

speed(3000)
penup()
goto(-140, 140)

# racing track
heading1 = Turtle()
heading1.penup()
heading1.goto(0, 180)
text = "WHAT'S FOR LUNCH!"
font_size = 16
font_style = ("Arial", font_size, "normal")
heading1.write(text, align="center", font=font_style)
heading1.hideturtle()
speed(0)
delay(0)
for step in range(15):
    write(step, align='center')
    right(90)

    for num in range(8):
        penup()
        forward(12)
        pendown()
        forward(12)
    penup()
    backward(12 * 2 * 8)
    left(90)
    forward(20)
delay(1)
speed(1)
horse = "horse.gif"
turtle.addshape(horse)

turtle_home=[]
for i in range(number_restaurant):
    player_1 = Turtle()
    player_1.color('red')
    player_1.shape(horse)
    player_1.penup()
    player_1.goto(-160, 110-40*i)
    player_1.pendown()
    player_1.shapesize(0.2, 0.2, 0.2)
    heading2 = Turtle()
    heading2.penup()
    heading2.goto(200, 110-40*i)
    text = restaurant[i]
    font_size = 12
    font_style = ("Arial", font_size, "normal")
    heading2.write(text, align="center", font=font_style)
    heading2.hideturtle()
    turtle_home.append(player_1)

layout = [
    [sg.Text('所有馬匹已經入閘 ')],
    [sg.Button('Start Now !'), sg.Button('Stop the game')]
]
window2 = sg.Window('Two ListBoxes Example', layout).read(close=True)
if window2[0] =="Stop the game":
    sys.exit()
time.sleep(2)
play_pos = []
while True:
    player_pos = [0,0,0,0]
    for i in range(number_restaurant):
        turtle_home[i].forward(randint(2, 8))
        player_pos[i]=turtle_home[i].pos()[0]

    if max(player_pos) > 140:
        winner = player_pos.index(max(player_pos))
        time.sleep(2)
        psg.popup(restaurant[winner] + "!")
        if restaurant[winner] == "TamJai/三哥":
            for i in turtles():
                i.reset()
                i.clear()
                i.hideturtle()

        else:
            bye()
        break

if restaurant[winner] == "TamJai/三哥":
    layout = [
        [sg.Text('Ready for game 2?')],
        [sg.Button('Yes'),sg.Button('Stop the game')]
    ]
    window3 = sg.Window('Two ListBoxes Example', layout).read(close=True)
    if window3[0]=="Stop the game":
        sys.exit()
    else:
        sc = turtle.Screen()
        speed(3000)
        heading = Turtle()
        penup()
        goto(-140, 140)

        # racing track
        heading.penup()
        heading.goto(0, 180)
        text =  "TamJai VS 三哥"
        font_size = 16
        font_style = ("Arial", font_size, "normal")
        heading.write(text, align="center", font=font_style)
        heading.hideturtle()
        speed(0)
        delay(0)
        for step in range(15):
            write(step, align='center')
            right(90)
            for num in range(8):
                penup()
                forward(12)
                pendown()
                forward(12)

            penup()
            backward(12 * 2 * 8)
            left(90)
            forward(20)
        delay(1)
        speed(1)
        player_2=Turtle()
        player_2.penup()
        player_2.showturtle()
        player_2.goto(-160, 70)
        player_2.pendown()
        player_2.shapesize(0.2, 0.2, 0.2)
        heading = Turtle()
        heading.penup()
        heading.goto(200, 70)
        text = "TamJai"
        font_size = 12
        font_style = ("Arial", font_size, "normal")
        heading.write(text, align="center", font=font_style)
        heading.hideturtle()
        player_3 = Turtle()
        player_3.showturtle()
        player_3.penup()
        player_3.goto(-160, 30)
        player_3.pendown()
        player_3.shapesize(0.2, 0.2, 0.2)

        heading = Turtle()
        heading.penup()
        heading.goto(200, 30)
        text = "三哥"
        font_size = 12
        font_style = ("Arial", font_size, "normal")
        heading.write(text, align="center", font=font_style)
        heading.hideturtle()
        while True:
            player_pos = []
            player_2.forward(randint(2, 8))
            player_3.forward(randint(2, 8))
            player_pos.append(player_2.pos()[0])
            player_pos.append(player_3.pos()[0])
            if max(player_pos) > 140:
                psg.popup(("TamJai" if player_pos.index(max(player_pos)) ==0 else "三哥") + "!")
                break
#update "上次食係幾時" in excel
wb = load_workbook("restaurant_list.xlsx")
ws = wb[window[1]["-LISTBOX1-"][0]]
monday_thisweek = datetime.today() - timedelta(datetime.today().weekday())
for i in range(2,ws.max_row+1):
    if ws["A"+str(i+1)].value == restaurant[winner]:
        ws["D" +str(i+1)].value  = '=Date('+str(datetime.today().year)+","+str(datetime.today().month)+","+str(datetime.today().day)+")"
        ws["D" + str(i+1)].number_format = "yyyy/mm/dd"
        break

wb.save("restaurant_list.xlsx")

os.remove('horse.gif')

